from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import and_, or_, select
from typing import List, Optional
import csv
import json
from datetime import datetime
from io import StringIO
import openpyxl
from io import BytesIO

from database import get_db
from models import Book, User, BookStatus, Transaction, Reservation
from routers.auth import get_current_user
from pydantic import BaseModel, ConfigDict

router = APIRouter()

@router.post("/reserve/{book_id}")
async def reserve_book(
    book_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Book).where(Book.id == book_id))
    book = result.scalar_one_or_none()
    if not book:
        raise HTTPException(status_code=404, detail="book not found")
    if book.status == BookStatus.RESERVED:
        raise HTTPException(status_code=400, detail="book already reserved")
    book.status = BookStatus.RESERVED
    await db.commit()
    await db.refresh(book)
    return {"message": "book reserved"}

class BookCreate(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    isbn: Optional[str] = None
    title: str
    author: Optional[str] = None
    tags: Optional[str] = None  # comma-separated tags like "fiction, romance, historical"
    description: Optional[str] = None
    price: float
    stock: int = 1
    status: BookStatus = BookStatus.IN_STOCK
    is_for_sale: bool = True
    is_for_rent: bool = False
    weekly_fee: Optional[float] = None
    condition: Optional[str] = None

class BookResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    isbn: Optional[str] = None
    title: str
    author: Optional[str] = None
    tags: Optional[str] = None  # Comma-separated tags
    description: Optional[str] = None
    price: float
    stock: int
    status: BookStatus
    is_for_sale: bool
    is_for_rent: bool
    weekly_fee: Optional[float] = None
    condition: Optional[str] = None
    owner_id: int

class BookUpdate(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    isbn: Optional[str] = None
    title: Optional[str] = None
    author: Optional[str] = None
    tags: Optional[str] = None  # Comma-separated tags
    description: Optional[str] = None
    price: Optional[float] = None
    stock: Optional[int] = None
    status: Optional[BookStatus] = None
    is_for_sale: Optional[bool] = None
    is_for_rent: Optional[bool] = None
    weekly_fee: Optional[float] = None
    condition: Optional[str] = None

@router.post("/", response_model=BookResponse)
async def create_book(
    book: BookCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    
    # create search text by combining title, author, and tags
    search_text = f"{book.title} {book.author or ''} {book.tags or ''} {book.description or ''}".strip()
    
    # Use model_dump() for Pydantic v2 or dict() for v1
    try:
        book_data = book.model_dump()  # Pydantic v2
    except AttributeError:
        book_data = book.dict()  # Pydantic v1
        
    book_data['search_text'] = search_text
    book_data['owner_id'] = current_user.id
    
    db_book = Book(**book_data)
    db.add(db_book)
    await db.commit()
    await db.refresh(db_book)
    return db_book

@router.get("/", response_model=List[BookResponse])
async def search_books(
    q: str = None,
    author: str = None,
    tags: str = None,  # Search within comma-separated tags
    min_price: float = None,
    max_price: float = None,
    city: str = None,
    for_sale: bool = None,
    for_rent: bool = None,
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db)
):
    query = select(Book).join(User)
    
    filters = []
    
    if q:
        filters.append(
            or_(
                Book.title.ilike(f"%{q}%"),
                Book.author.ilike(f"%{q}%"),
                Book.description.ilike(f"%{q}%")
            )
        )
    
    if author:
        filters.append(Book.author.ilike(f"%{author}%"))
    
    if tags:
        filters.append(Book.tags.ilike(f"%{tags}%"))
    
    if min_price is not None:
        filters.append(Book.price >= min_price)
    
    if max_price is not None:
        filters.append(Book.price <= max_price)
    
    if city:
        filters.append(User.city.ilike(f"%{city}%"))
    
    if for_sale is not None:
        filters.append(Book.is_for_sale == for_sale)
    
    if for_rent is not None:
        filters.append(Book.is_for_rent == for_rent)
    
    if filters:
        query = query.where(and_(*filters))
    
    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    books = result.scalars().all()
    return books

@router.get("/reservations")
async def get_user_reservations(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get all reservations made by the current user"""
    from models import Reservation, ReservationStatus
    
    query = select(
        Reservation,
        Book,
        User.first_name.label("seller_first_name"),
        User.last_name.label("seller_last_name"), 
        User.email.label("seller_email"),
        User.phone.label("seller_phone"),
        User.city.label("seller_city"),
        User.state.label("seller_state")
    ).join(
        Book, Reservation.book_id == Book.id
    ).join(
        User, Book.owner_id == User.id
    ).where(
        Reservation.user_id == current_user.id
    ).order_by(Reservation.created_at.desc())
    
    result = await db.execute(query)
    reservations_data = result.all()
    
    reservations = []
    for reservation, book, seller_first_name, seller_last_name, seller_email, seller_phone, seller_city, seller_state in reservations_data:
        # Create location string from seller's location
        book_location = ""
        if seller_city and seller_state:
            book_location = f"{seller_city}, {seller_state}"
        elif seller_city:
            book_location = seller_city
        elif seller_state:
            book_location = seller_state
        else:
            book_location = "Location not specified"
            
        reservation_dict = {
            "id": reservation.id,
            "book_title": book.title,
            "book_author": book.author or "Unknown",
            "total_price": float(book.price),
            "amount_paid": float(reservation.reservation_fee),
            "remaining_amount": float(book.price - reservation.reservation_fee),
            "status": reservation.status.value,
            "created_at": reservation.created_at.isoformat() if reservation.created_at else None,
            "valid_until": reservation.expires_at.isoformat() if reservation.expires_at else None,
            "book_location": book_location
        }
        
        # Only show seller contact if payment is confirmed
        if reservation.status == ReservationStatus.CONFIRMED:
            reservation_dict.update({
                "seller_contact": True,
                "seller_name": f"{seller_first_name} {seller_last_name}",
                "seller_email": seller_email,
                "seller_phone": seller_phone
            })
        else:
            reservation_dict["seller_contact"] = False
    
        reservations.append(reservation_dict)
    
    return reservations

@router.get("/{book_id}", response_model=BookResponse)
async def get_book(book_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Book).where(Book.id == book_id))
    book = result.scalar_one_or_none()
    if not book:
        raise HTTPException(status_code=404, detail="book not found")
    return book

@router.put("/{book_id}", response_model=BookResponse)
async def update_book(
    book_id: int,
    book_update: BookUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Book).where(Book.id == book_id))
    book = result.scalar_one_or_none()
    if not book:
        raise HTTPException(status_code=404, detail="book not found")
    
    if book.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="not authorized to update this book")
    
    update_data = book_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(book, field, value)
    
    await db.commit()
    await db.refresh(book)
    return book

@router.delete("/{book_id}")
async def delete_book(
    book_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Book).where(Book.id == book_id))
    book = result.scalar_one_or_none()
    if not book:
        raise HTTPException(status_code=404, detail="book not found")
    
    if book.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="not authorized to delete this book")
    
    await db.delete(book)
    await db.commit()
    return {"message": "book deleted"}

@router.post("/import")
async def import_books(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    
    if not file.filename.endswith(('.csv', '.xlsx')):
        raise HTTPException(status_code=400, detail="only csv and xlsx files supported")
    
    try:
        books_created = 0
        
        if file.filename.endswith('.csv'):
            # handle csv files
            content = await file.read()
            csv_content = content.decode('utf-8')
            csv_reader = csv.DictReader(StringIO(csv_content))
            
            required_columns = ['title', 'author', 'price']
            headers = csv_reader.fieldnames
            if not all(col in headers for col in required_columns):
                raise HTTPException(
                    status_code=400, 
                    detail=f"missing required columns: {required_columns}"
                )
            
            for row in csv_reader:
                book_data = {
                    'title': row['title'],
                    'author': row['author'],
                    'price': float(row['price']),
                    'owner_id': current_user.id
                }
                
                # optional fields
                if 'isbn' in row and row['isbn']:
                    book_data['isbn'] = str(row['isbn'])
                if 'tags' in row and row['tags']:
                    book_data['tags'] = row['tags']  # Should be comma-separated
                if 'description' in row and row['description']:
                    book_data['description'] = row['description']
                if 'stock' in row and row['stock']:
                    book_data['stock'] = int(row['stock'])
                if 'condition' in row and row['condition']:
                    book_data['condition'] = row['condition']
                
                db_book = Book(**book_data)
                db.add(db_book)
                books_created += 1
        
        else:
            # handle xlsx files
            content = await file.read()
            wb = openpyxl.load_workbook(filename=content, read_only=True)
            ws = wb.active
            
            # get headers from first row
            headers = [cell.value for cell in ws[1]]
            required_columns = ['title', 'author', 'price']
            
            if not all(col in headers for col in required_columns):
                raise HTTPException(
                    status_code=400, 
                    detail=f"missing required columns: {required_columns}"
                )
            
            # create column index mapping
            col_indices = {header: idx for idx, header in enumerate(headers)}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[col_indices['title']]:  # skip empty rows
                    continue
                    
                book_data = {
                    'title': row[col_indices['title']],
                    'author': row[col_indices['author']],
                    'price': float(row[col_indices['price']]),
                    'owner_id': current_user.id
                }
                
                # optional fields
                if 'isbn' in col_indices and row[col_indices['isbn']]:
                    book_data['isbn'] = str(row[col_indices['isbn']])
                if 'tags' in col_indices and row[col_indices['tags']]:
                    book_data['tags'] = row[col_indices['tags']]  # Should be comma-separated
                if 'description' in col_indices and row[col_indices['description']]:
                    book_data['description'] = row[col_indices['description']]
                if 'stock' in col_indices and row[col_indices['stock']]:
                    book_data['stock'] = int(row[col_indices['stock']])
                if 'condition' in col_indices and row[col_indices['condition']]:
                    book_data['condition'] = row[col_indices['condition']]
                
                db_book = Book(**book_data)
                db.add(db_book)
                books_created += 1
        
        await db.commit()
        return {"message": f"imported {books_created} books successfully"}
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"import failed: {str(e)}")

@router.get("/my/books", response_model=List[BookResponse])
async def get_my_books(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Book).filter(Book.owner_id == current_user.id))
    books = result.scalars().all()
    return books

class ReservationInfo(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    reservation_id: int
    buyer_id: int
    buyer_name: str
    buyer_email: str
    reservation_fee: float
    status: str
    created_at: datetime

class BookWithReservationResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    
    id: int
    isbn: Optional[str] = None
    title: str
    author: Optional[str] = None
    tags: Optional[str] = None
    description: Optional[str] = None
    price: float
    stock: int
    status: BookStatus
    is_for_sale: bool
    is_for_rent: bool
    weekly_fee: Optional[float] = None
    condition: Optional[str] = None
    owner_id: int
    created_at: Optional[datetime] = None
    reservation: Optional[ReservationInfo] = None

@router.get("/my/books/with-reservations", response_model=List[BookWithReservationResponse])
async def get_my_books_with_reservations(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    # Get books owned by the current user with their reservations
    result = await db.execute(
        select(Book)
        .filter(Book.owner_id == current_user.id)
        .order_by(Book.created_at.desc())
    )
    books = result.scalars().all()
    
    books_with_reservations = []
    for book in books:
        # Get the most recent active reservation for this book
        reservation_result = await db.execute(
            select(Reservation)
            .filter(Reservation.book_id == book.id)
            .order_by(Reservation.created_at.desc())
        )
        reservation = reservation_result.scalars().first()  # Use first() instead of scalar_one_or_none()
        
        reservation_info = None
        if reservation:
            # Get buyer information
            buyer_result = await db.execute(
                select(User).filter(User.id == reservation.user_id)
            )
            buyer = buyer_result.scalar_one_or_none()
            
            if buyer:
                reservation_info = ReservationInfo(
                    reservation_id=reservation.id,
                    buyer_id=buyer.id,
                    buyer_name=f"{buyer.first_name} {buyer.last_name}",
                    buyer_email=buyer.email,
                    reservation_fee=reservation.reservation_fee,
                    status=reservation.status.value,
                    created_at=reservation.created_at
                )
        
        book_dict = {
            "id": book.id,
            "isbn": book.isbn,
            "title": book.title,
            "author": book.author,
            "tags": book.tags,
            "description": book.description,
            "price": book.price,
            "stock": book.stock,
            "status": book.status,
            "is_for_sale": book.is_for_sale,
            "is_for_rent": book.is_for_rent,
            "weekly_fee": book.weekly_fee,
            "condition": book.condition,
            "owner_id": book.owner_id,
            "created_at": book.created_at,
            "reservation": reservation_info
        }
        books_with_reservations.append(book_dict)
    
    return books_with_reservations


@router.post('/import-excel')
async def import_books_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Import books from an uploaded Excel file. Expected header columns:
       title, author, price, stock, is_for_sale, is_for_rent, weekly_fee, condition, tags, description, isbn
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file must contain a header row and at least one data row")

    header = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    # Map header names to indices
    idx = {name: i for i, name in enumerate(header)}

    created = 0
    errors = []

    for rnum, row in enumerate(rows[1:], start=2):
        try:
            title = row[idx.get('title')] if 'title' in idx else None
            if not title:
                raise ValueError('Missing title')

            author = row[idx.get('author')] if 'author' in idx else None
            price = row[idx.get('price')] if 'price' in idx else None
            stock = row[idx.get('stock')] if 'stock' in idx else 1
            is_for_sale = row[idx.get('is_for_sale')] if 'is_for_sale' in idx else True
            is_for_rent = row[idx.get('is_for_rent')] if 'is_for_rent' in idx else False
            weekly_fee = row[idx.get('weekly_fee')] if 'weekly_fee' in idx else None
            condition = row[idx.get('condition')] if 'condition' in idx else None
            tags = row[idx.get('tags')] if 'tags' in idx else None
            description = row[idx.get('description')] if 'description' in idx else None
            isbn = row[idx.get('isbn')] if 'isbn' in idx else None

            # Basic parsing
            try:
                price = float(price)
            except Exception:
                raise ValueError('Invalid price')

            try:
                stock = int(stock)
            except Exception:
                stock = 1

            book = Book(
                isbn=str(isbn) if isbn is not None else None,
                title=str(title),
                author=str(author) if author is not None else None,
                search_text=f"{title} {author or ''} {tags or ''} {description or ''}",
                description=str(description) if description is not None else None,
                tags=str(tags) if tags is not None else None,
                price=price,
                stock=stock,
                owner_id=current_user.id,
                is_for_sale=bool(is_for_sale),
                is_for_rent=bool(is_for_rent),
                weekly_fee=float(weekly_fee) if weekly_fee not in (None, '') else None,
                condition=str(condition) if condition is not None else None,
            )

            db.add(book)
            await db.commit()
            await db.refresh(book)
            created += 1
        except Exception as e:
            errors.append({'row': rnum, 'error': str(e)})

    return {"created": created, "errors": errors}


class TransactionResponse(BaseModel):
    id: int
    book_title: str
    buyer_id: int
    buyer_name: str
    price: float
    transaction_type: str
    status: str
    created_at: datetime

@router.get("/my/sales")
async def get_my_sales(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    # get all sales by current user
    result = await db.execute(
        select(Transaction, Book, User)
        .join(Book, Transaction.book_id == Book.id)
        .join(User, Transaction.buyer_id == User.id)
        .where(Transaction.seller_id == current_user.id)
        .order_by(Transaction.created_at.desc())
    )
    sales = result.all()
    
    return [
        {
            "id": transaction.id,
            "book_title": book.title,
            "buyer_id": buyer.id,
            "buyer_name": f"{buyer.first_name} {buyer.last_name}",
            "price": transaction.price,
            "transaction_type": transaction.transaction_type,
            "status": transaction.status,
            "created_at": transaction.created_at
        }
        for transaction, book, buyer in sales
    ]
