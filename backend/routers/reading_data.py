from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.orm import Session
from typing import List, Dict
import csv
import json
from datetime import datetime
from io import StringIO
import openpyxl

from database import get_db
from models import ReadingData, User
from routers.auth import get_current_user
from pydantic import BaseModel

router = APIRouter()

class ReadingDataCreate(BaseModel):
    book_isbn: str = None
    book_title: str
    author: str
    genre: str = None
    rating: int = None
    review: str = None
    date_read: datetime = None

class ReadingDataResponse(BaseModel):
    id: int
    book_isbn: str = None
    book_title: str
    author: str
    genre: str = None
    rating: int = None
    review: str = None
    date_read: datetime = None
    created_at: datetime

class TasteVector(BaseModel):
    genres: Dict[str, float]
    authors: Dict[str, float]
    avg_rating: float
    total_books: int

@router.post("/", response_model=ReadingDataResponse)
async def add_reading_data(
    reading_data: ReadingDataCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    db_reading_data = ReadingData(
        **reading_data.dict(),
        user_id=current_user.id
    )
    
    db.add(db_reading_data)
    db.commit()
    db.refresh(db_reading_data)
    
    return db_reading_data

@router.get("/", response_model=List[ReadingDataResponse])
async def get_my_reading_data(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    reading_data = db.query(ReadingData).filter(
        ReadingData.user_id == current_user.id
    ).order_by(ReadingData.date_read.desc()).all()
    
    return reading_data

@router.post("/import")
async def import_reading_data(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith(('.csv', '.xlsx')):
        raise HTTPException(status_code=400, detail="only csv and xlsx files supported")
    
    try:
        imported_count = 0
        
        if file.filename.endswith('.csv'):
            # handle csv files
            content = await file.read()
            csv_content = content.decode('utf-8')
            csv_reader = csv.DictReader(StringIO(csv_content))
            
            required_columns = ['book_title', 'author']
            headers = csv_reader.fieldnames
            if not all(col in headers for col in required_columns):
                raise HTTPException(
                    status_code=400, 
                    detail=f"missing required columns: {required_columns}"
                )
            
            for row in csv_reader:
                reading_data = {
                    'book_title': row['book_title'],
                    'author': row['author'],
                    'user_id': current_user.id
                }
                
                # optional fields
                if 'book_isbn' in row and row['book_isbn']:
                    reading_data['book_isbn'] = str(row['book_isbn'])
                if 'genre' in row and row['genre']:
                    reading_data['genre'] = row['genre']
                if 'rating' in row and row['rating']:
                    reading_data['rating'] = int(row['rating'])
                if 'review' in row and row['review']:
                    reading_data['review'] = row['review']
                if 'date_read' in row and row['date_read']:
                    try:
                        reading_data['date_read'] = datetime.strptime(row['date_read'], '%Y-%m-%d')
                    except ValueError:
                        pass  # skip invalid dates
                
                db_reading_data = ReadingData(**reading_data)
                db.add(db_reading_data)
                imported_count += 1
        
        else:
            # handle xlsx files
            content = await file.read()
            wb = openpyxl.load_workbook(filename=content, read_only=True)
            ws = wb.active
            
            # get headers from first row
            headers = [cell.value for cell in ws[1]]
            required_columns = ['book_title', 'author']
            
            if not all(col in headers for col in required_columns):
                raise HTTPException(
                    status_code=400, 
                    detail=f"missing required columns: {required_columns}"
                )
            
            # create column index mapping
            col_indices = {header: idx for idx, header in enumerate(headers)}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[col_indices['book_title']]:  # skip empty rows
                    continue
                    
                reading_data = {
                    'book_title': row[col_indices['book_title']],
                    'author': row[col_indices['author']],
                    'user_id': current_user.id
                }
                
                # optional fields
                if 'book_isbn' in col_indices and row[col_indices['book_isbn']]:
                    reading_data['book_isbn'] = str(row[col_indices['book_isbn']])
                if 'genre' in col_indices and row[col_indices['genre']]:
                    reading_data['genre'] = row[col_indices['genre']]
                if 'rating' in col_indices and row[col_indices['rating']]:
                    reading_data['rating'] = int(row[col_indices['rating']])
                if 'review' in col_indices and row[col_indices['review']]:
                    reading_data['review'] = row[col_indices['review']]
                if 'date_read' in col_indices and row[col_indices['date_read']]:
                    if isinstance(row[col_indices['date_read']], datetime):
                        reading_data['date_read'] = row[col_indices['date_read']]
                    else:
                        try:
                            reading_data['date_read'] = datetime.strptime(str(row[col_indices['date_read']]), '%Y-%m-%d')
                        except ValueError:
                            pass  # skip invalid dates
                
                db_reading_data = ReadingData(**reading_data)
                db.add(db_reading_data)
                imported_count += 1
        
        db.commit()
        return {"message": f"imported {imported_count} reading records successfully"}
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"import failed: {str(e)}")

@router.get("/taste-vector", response_model=TasteVector)
async def get_taste_vector(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    reading_data = db.query(ReadingData).filter(
        ReadingData.user_id == current_user.id
    ).all()
    
    if not reading_data:
        return TasteVector(
            genres={},
            authors={},
            avg_rating=0.0,
            total_books=0
        )
    
    # calculate genre preferences
    genre_counts = {}
    author_counts = {}
    ratings = []
    
    for record in reading_data:
        if record.genre:
            genre_counts[record.genre] = genre_counts.get(record.genre, 0) + 1
        
        author_counts[record.author] = author_counts.get(record.author, 0) + 1
        
        if record.rating:
            ratings.append(record.rating)
    
    total_books = len(reading_data)
    
    # normalize genre preferences
    genres = {genre: count / total_books for genre, count in genre_counts.items()}
    authors = {author: count / total_books for author, count in author_counts.items()}
    avg_rating = sum(ratings) / len(ratings) if ratings else 0.0
    
    return TasteVector(
        genres=genres,
        authors=authors,
        avg_rating=avg_rating,
        total_books=total_books
    )

@router.get("/recommendations")
async def get_recommendations(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # get user's taste vector
    taste_vector = await get_taste_vector(current_user, db)
    
    # find users with similar taste (simplified collaborative filtering)
    similar_users = []
    all_users = db.query(User).filter(User.id != current_user.id).all()
    
    for user in all_users:
        user_reading_data = db.query(ReadingData).filter(
            ReadingData.user_id == user.id
        ).all()
        
        if not user_reading_data:
            continue
        
        # calculate similarity score based on common genres
        user_genres = {}
        for record in user_reading_data:
            if record.genre:
                user_genres[record.genre] = user_genres.get(record.genre, 0) + 1
        
        total_user_books = len(user_reading_data)
        user_genre_prefs = {genre: count / total_user_books for genre, count in user_genres.items()}
        
        # calculate cosine similarity
        common_genres = set(taste_vector.genres.keys()) & set(user_genre_prefs.keys())
        if common_genres:
            similarity = sum(
                taste_vector.genres[genre] * user_genre_prefs[genre] 
                for genre in common_genres
            )
            similar_users.append((user.id, similarity))
    
    # sort by similarity and get top 5
    similar_users.sort(key=lambda x: x[1], reverse=True)
    top_similar_users = [user_id for user_id, _ in similar_users[:5]]
    
    return {
        "taste_vector": taste_vector,
        "similar_users": top_similar_users,
        "club_suggestions": f"consider joining clubs for genres: {', '.join(list(taste_vector.genres.keys())[:3])}"
    }
